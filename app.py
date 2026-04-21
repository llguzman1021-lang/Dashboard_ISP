import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import bcrypt, math, pytz, calendar, re
from sqlalchemy import create_engine, text
from datetime import datetime, timedelta, date, time as datetime_time
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.units import cm

# =====================================================================
# CONFIGURACIÓN GLOBAL
# =====================================================================
st.set_page_config(
    page_title="Multinet NOC",
    layout="wide",
    page_icon="📊",
    initial_sidebar_state="collapsed",
)
SV_TZ = pytz.timezone('America/El_Salvador')

COLOR_PRIMARY   = '#f15c22'
COLOR_SECONDARY = '#1d2c59'
COLOR_TEAL      = '#29b09d'
COLOR_DANGER    = '#ff2b2b'
PALETA_CORP     = (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_TEAL, '#ff9f43', '#83c9ff', COLOR_DANGER)
SUPERADMIN_USERNAME = 'Admin'

DEFAULT_ZONAS = (
    ("El Rosario", 13.4886, -89.0256), ("ARG", 13.4880, -89.3200), ("Tepezontes", 13.6214, -89.0125),
    ("La Libertad", 13.4883, -89.3200), ("El Tunco", 13.4930, -89.3830), ("Costa del Sol", 13.3039, -88.9450),
    ("Zacatecoluca", 13.5048, -88.8710), ("Zaragoza", 13.5850, -89.2890), ("Santiago Nonualco", 13.5186, -88.9442),
    ("Rio Mar", 13.4900, -89.3500), ("San Salvador (Central)", 13.6929, -89.2182),
)
DEFAULT_EQUIPOS = (
    "ONT", "Repetidor Wi-Fi", "Antena Ubiquiti", "OLT", "Caja NAP", "Switch", "Fibra Principal",
    "Servidor de Aplicativos", "Servidor DNS", "Servidor Virtual", "Mikrotik Concentrador",
    "Encoder", "Caja Señal TV", "Antena Señal TV"
)
DEFAULT_CAUSAS = (
    ("Corte de Fibra por Terceros", True), ("Corte de Fibra (No Especificado)", False), ("Caída de Árboles sobre Fibra", True),
    ("Falla de Energía Comercial", True), ("Corrosión en Equipos", False), ("Daños por Fauna", True),
    ("Falla de Hardware", False), ("Falla de Configuración", False), ("Falla de Redundancia", False),
    ("Saturación de Tráfico", False), ("Saturación en Servidor", False), ("Mantenimiento Programado", False),
    ("Vandalismo o Hurto", True), ("Condiciones Climáticas", True),
)
DEFAULT_SERVICIOS  = ("Internet", "Cable TV (CATV)", "IPTV (Mnet+)", "Internet/Cable TV", "Aplicativos internos")
DEFAULT_CATEGORIAS = ("Red Multinet", "Cliente Corporativo", "Falla Interna (No afecta clientes)")
CAT_INTERNA = "Falla Interna (No afecta clientes)"
MESES = ("Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre")
_ALLOWED_TABLES = {"cat_zonas", "cat_equipos", "cat_causas", "cat_servicios", "cmdb_nodos"}

st.markdown("""
<style>
div.stButton > button { border: none !important; outline: none !important; box-shadow: none !important; border-radius: 8px; width: 100%; font-weight: 600; transition: all 0.3s ease !important; }
div.stButton > button:hover  { transform: translateY(-2px); }
[data-testid="stMetricValue"] { color: #ffffff !important; font-size: 34px !important; font-weight: 800 !important; }
[data-testid="stMetricLabel"] { color: #a5a8b5 !important; font-size: 14px !important; font-weight: 500 !important; }
[data-testid="stMetricDelta"] svg { width: 20px; height: 20px; }
[data-testid="stMetricDelta"] div { font-size: 13px !important; font-weight: 600 !important; }
button[data-baseweb="tab"] { background-color: #1e1e2f !important; border-radius: 12px 12px 0 0 !important; margin-right: 8px !important; padding: 13px 26px !important; border: 2px solid #333 !important; border-bottom: none !important; transition: all 0.3s; }
button[data-baseweb="tab"]:hover { background-color: #2a2a3f !important; }
button[data-baseweb="tab"][aria-selected="true"] { background-color: #f15c22 !important; border-color: #f15c22 !important; }
button[data-baseweb="tab"] p { font-size: 15px !important; font-weight: 700 !important; color: #a5a8b5 !important; margin: 0 !important; }
button[data-baseweb="tab"][aria-selected="true"] p { color: #ffffff !important; }
.st-emotion-cache-1wivap2 { padding-top: 1.5rem; }
</style>
""", unsafe_allow_html=True)

# ── Inicialización SEGURA de sesión ──
_sesion_defaults = {
    'form_reset': 0, 'logged_in': False, 'role': '', 'username': '',
    'log_u': '', 'log_p': '', 'log_err': '', 'flash_msg': '', 'flash_type': '',
    'data_version': 0, 'pdf_cache_key': '', 'pdf_cache_bytes': b'',
}
for _k, _v in _sesion_defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

if st.session_state.flash_msg:
    if st.session_state.flash_type == 'error':
        st.error(st.session_state.flash_msg, icon="❌")
    else:
        st.toast(st.session_state.flash_msg, icon="✅")
    st.session_state.flash_msg  = ""
    st.session_state.flash_type = ""

def _dv() -> int:
    return st.session_state.get('data_version', 0)

def _invalidar_cache():
    """Invalida solo datos, manteniendo caché de catálogos y otros meses."""
    st.session_state.data_version    = _dv() + 1
    st.session_state.pdf_cache_key   = ""   # también invalida el PDF cacheado

# =====================================================================
# BASE DE DATOS
# =====================================================================
@st.cache_resource
def get_engine():
    return create_engine(st.secrets["neon_dsn"], pool_pre_ping=True, pool_recycle=300)

engine = get_engine()

def hash_pw(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def check_pw(p: str, h: str) -> bool:
    return bcrypt.checkpw(p.encode(), h.encode())

def validar_password(p: str):
    if len(p) < 8:              return "Mínimo 8 caracteres."
    if not re.search(r'[A-Z]', p): return "Debe incluir mayúscula."
    if not re.search(r'[a-z]', p): return "Debe incluir minúscula."
    if not re.search(r'\d',    p): return "Debe incluir un número."
    if not re.search(r'[!@#$%^&*()_+\-=\[\]{};\':"\\|,.<>\/?`~]', p):
        return "Debe incluir un carácter especial (!@#$%...)."
    return None

def init_db():
    with engine.begin() as conn:
        conn.execute(text("CREATE TABLE IF NOT EXISTS users (id SERIAL PRIMARY KEY, username VARCHAR(50) UNIQUE, password_hash VARCHAR(255), role VARCHAR(20), failed_attempts INT DEFAULT 0, locked_until TIMESTAMP, is_banned BOOLEAN DEFAULT FALSE)"))
        conn.execute(text("CREATE TABLE IF NOT EXISTS audit_logs (id SERIAL PRIMARY KEY, timestamp TIMESTAMP, username VARCHAR(50), action VARCHAR(50), details TEXT)"))
        conn.execute(text("CREATE TABLE IF NOT EXISTS cat_zonas (id SERIAL PRIMARY KEY, nombre VARCHAR(150) UNIQUE NOT NULL, lat FLOAT DEFAULT 13.6929, lon FLOAT DEFAULT -89.2182)"))
        conn.execute(text("CREATE TABLE IF NOT EXISTS cat_equipos (id SERIAL PRIMARY KEY, nombre VARCHAR(100) UNIQUE NOT NULL)"))
        conn.execute(text("CREATE TABLE IF NOT EXISTS cat_causas  (id SERIAL PRIMARY KEY, nombre VARCHAR(150) UNIQUE NOT NULL, es_externa BOOLEAN DEFAULT FALSE)"))
        conn.execute(text("CREATE TABLE IF NOT EXISTS cat_servicios (id SERIAL PRIMARY KEY, nombre VARCHAR(100) UNIQUE NOT NULL)"))
        conn.execute(text("CREATE TABLE IF NOT EXISTS cmdb_nodos (id SERIAL PRIMARY KEY, zona VARCHAR(150), equipo VARCHAR(100), clientes INT DEFAULT 0, fuente VARCHAR(50) DEFAULT 'Manual', ultima_sincronizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP, UNIQUE(zona, equipo))"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS incidents (
                id SERIAL PRIMARY KEY, zona VARCHAR(150) NOT NULL, subzona VARCHAR(150),
                afectacion_general BOOLEAN DEFAULT TRUE, servicio VARCHAR(100), categoria VARCHAR(100),
                equipo_afectado VARCHAR(100), estado VARCHAR(20) DEFAULT 'Cerrado',
                inicio_incidente TIMESTAMPTZ NOT NULL, fin_incidente TIMESTAMPTZ,
                clientes_afectados INT DEFAULT 0, impacto_porcentaje FLOAT DEFAULT 0.0,
                causa_raiz VARCHAR(150), descripcion TEXT, duracion_horas FLOAT,
                conocimiento_tiempos VARCHAR(50) DEFAULT 'Total', deleted_at TIMESTAMPTZ DEFAULT NULL
            )"""))
        if conn.execute(text("SELECT count(*) FROM cat_zonas")).scalar() == 0:
            for n, la, lo in DEFAULT_ZONAS:
                conn.execute(text("INSERT INTO cat_zonas (nombre,lat,lon) VALUES (:n,:lat,:lon) ON CONFLICT DO NOTHING"), {"n": n, "lat": la, "lon": lo})
        if conn.execute(text("SELECT count(*) FROM cat_equipos")).scalar() == 0:
            for e in DEFAULT_EQUIPOS:
                conn.execute(text("INSERT INTO cat_equipos (nombre) VALUES (:n) ON CONFLICT DO NOTHING"), {"n": e})
        if conn.execute(text("SELECT count(*) FROM cat_causas")).scalar() == 0:
            for n, ext in DEFAULT_CAUSAS:
                conn.execute(text("INSERT INTO cat_causas (nombre,es_externa) VALUES (:n,:e) ON CONFLICT DO NOTHING"), {"n": n, "e": ext})
        if conn.execute(text("SELECT count(*) FROM cat_servicios")).scalar() == 0:
            for s in DEFAULT_SERVICIOS:
                conn.execute(text("INSERT INTO cat_servicios (nombre) VALUES (:n) ON CONFLICT DO NOTHING"), {"n": s})
        # FIX #1: Query parametrizada — elimina riesgo de SQL injection en init_db
        if conn.execute(text("SELECT count(*) FROM users WHERE username=:u"), {"u": SUPERADMIN_USERNAME}).scalar() == 0:
            conn.execute(text("INSERT INTO users (username,password_hash,role) VALUES (:u,:h,'admin')"),
                         {"u": SUPERADMIN_USERNAME, "h": hash_pw("Areakde5@")})

try:
    init_db()
except Exception as _e:
    st.error(f"Error DB Inicialización: {_e}")

@st.cache_data(ttl=300, show_spinner=False)
def get_zonas() -> list:
    try:
        with engine.connect() as c:
            return [(r[0], r[1], r[2]) for r in c.execute(text("SELECT nombre, lat, lon FROM cat_zonas ORDER BY nombre")).fetchall()]
    except Exception:
        return list(DEFAULT_ZONAS)

@st.cache_data(ttl=300, show_spinner=False)
def get_cat(tabla: str) -> list:
    if tabla not in _ALLOWED_TABLES: return []
    try:
        with engine.connect() as c:
            return [r[0] for r in c.execute(text(f"SELECT nombre FROM {tabla} ORDER BY nombre")).fetchall()]
    except Exception:
        return []

@st.cache_data(ttl=300, show_spinner=False)
def get_causas_con_flag() -> dict:
    try:
        with engine.connect() as c:
            return {r[0]: r[1] for r in c.execute(text("SELECT nombre, es_externa FROM cat_causas ORDER BY nombre")).fetchall()}
    except Exception:
        return {}

def clear_catalog_cache():
    get_zonas.clear(); get_cat.clear(); get_causas_con_flag.clear()

@st.cache_data(ttl=300, show_spinner=False)
def get_all_cmdb_nodos(cache_version: int = 0) -> dict:
    try:
        with engine.connect() as conn:
            res = conn.execute(text("SELECT zona, equipo, clientes FROM cmdb_nodos")).fetchall()
            return {(str(r[0]).lower(), str(r[1]).lower()): r[2] for r in res}
    except Exception:
        return {}

def get_clientes_cmdb(zona: str, equipo: str) -> int:
    if zona == "San Salvador (Central)": return 0
    zona_l = zona.lower(); equipo_l = equipo.lower()
    for (z, e), cl in get_all_cmdb_nodos(cache_version=_dv()).items():
        if z == zona_l and e == equipo_l:
            return cl
    return 0

# =====================================================================
# CARGA + ENRIQUECIMIENTO — COMBINADOS Y CACHEADOS
# FIX #2: enriquecer ya no se llama fuera del caché.
# Se usa _load_raw (sin caché) + _enriquecer (sin caché) internamente,
# y load_and_enrich (con caché) como punto de entrada único.
# FIX #3: tz_convert se aplica a la columna completa (no con .loc)
# para evitar ValueError en pandas ≥ 2.0.
# =====================================================================
def _load_raw(fecha_ini: date, fecha_fin: date, include_deleted: bool,
              zona_filtro: str, serv_filtro: str, seg_filtro: str) -> pd.DataFrame:
    s_date = datetime.combine(fecha_ini, datetime_time(0, 0, 0))
    e_date = datetime.combine(fecha_fin, datetime_time(23, 59, 59))
    params: dict = {"s": s_date, "e": e_date}
    # FIX #8: la 4ª condición incluye lower bound para tickets abiertos,
    # evitando traer tickets de años anteriores al periodo actual.
    year_start = datetime(fecha_ini.year, 1, 1)
    params["year_s"] = year_start
    conds = [
        "(  (inicio_incidente >= :s AND inicio_incidente <= :e)"
        "  OR (fin_incidente   >= :s AND fin_incidente   <= :e)"
        "  OR (inicio_incidente <= :s AND fin_incidente  >= :e)"
        "  OR (inicio_incidente >= :year_s AND inicio_incidente <= :e AND estado = 'Abierto')"
        ")"
    ]
    conds.append("deleted_at IS NULL" if not include_deleted else "deleted_at IS NOT NULL")
    if zona_filtro != "Todas":
        conds.append("zona = :zona_f");  params["zona_f"] = zona_filtro
    if serv_filtro != "Todos":
        conds.append("servicio = :serv_f"); params["serv_f"] = serv_filtro
    if seg_filtro  != "Todos":
        conds.append("categoria = :seg_f"); params["seg_f"] = seg_filtro
    q = "SELECT * FROM incidents WHERE " + " AND ".join(conds) + " ORDER BY inicio_incidente ASC"
    try:
        with engine.connect() as conn:
            return pd.read_sql(text(q), conn, params=params)
    except Exception:
        return pd.DataFrame()

def _enriquecer(df: pd.DataFrame) -> pd.DataFrame:
    """Procesamiento pandas puro, sin I/O. No necesita caché propia."""
    if df is None: return pd.DataFrame()
    df = df.copy()
    df.columns = [c.lower() for c in df.columns]

    df['duracion_horas']     = pd.to_numeric(df.get('duracion_horas',     0), errors='coerce').fillna(0.0)
    df['clientes_afectados'] = pd.to_numeric(df.get('clientes_afectados', 0), errors='coerce').fillna(0).astype(int)

    if df.empty:
        df['Severidad']     = pd.Series(dtype='object')
        df['es_externa']    = pd.Series(dtype='bool')
        df['zona_completa'] = pd.Series(dtype='object')
        return df

    # FIX #3: tz_convert sobre columna completa (no .loc parcial) — pandas 2.0 safe
    if 'inicio_incidente' in df.columns:
        df['inicio_incidente'] = pd.to_datetime(df['inicio_incidente'], errors='coerce', utc=True)
        df['inicio_incidente'] = df['inicio_incidente'].dt.tz_convert(SV_TZ)

    if 'fin_incidente' in df.columns:
        df['fin_incidente'] = pd.to_datetime(df['fin_incidente'], errors='coerce', utc=True)
        df['fin_incidente'] = df['fin_incidente'].dt.tz_convert(SV_TZ)

    causas_ext_map = get_causas_con_flag()

    def _sev(r) -> str:
        if r.get('estado') == 'Abierto':                                                return '🚨 CRÍTICA (En Curso)'
        if r.get('categoria') == CAT_INTERNA:                                           return '🟢 P4 (Interna)'
        if r.get('duracion_horas', 0) >= 12 or r.get('clientes_afectados', 0) >= 1000: return '🔴 P1 (Crítica)'
        if r.get('duracion_horas', 0) >= 4  or r.get('clientes_afectados', 0) >= 300:  return '🟠 P2 (Alta)'
        return '🟡 P3 (Media)'

    df['Severidad'] = df.apply(_sev, axis=1)
    # FIX #10: Series con index correcto cuando causa_raiz no existe
    df['es_externa'] = (
        df['causa_raiz'].map(lambda x: causas_ext_map.get(x, False))
        if 'causa_raiz' in df.columns
        else pd.Series(False, index=df.index)
    )
    df['zona_completa'] = df.apply(
        lambda r: f"{r.get('zona','')} (General)" if r.get('afectacion_general', True)
                  else f"{r.get('zona','')} - {r.get('subzona','')}",
        axis=1
    )
    return df

@st.cache_data(ttl=60, show_spinner=False)
def load_and_enrich(
    fecha_ini: date, fecha_fin: date, include_deleted: bool = False,
    zona_filtro: str = "Todas", serv_filtro: str = "Todos", seg_filtro: str = "Todos",
    cache_version: int = 0,
) -> pd.DataFrame:
    """Único punto de entrada: carga + enriquecimiento, todo cacheado juntos."""
    return _enriquecer(_load_raw(fecha_ini, fecha_fin, include_deleted, zona_filtro, serv_filtro, seg_filtro))

# ── Strip timezone a nivel de DataFrame completo (columna por columna) ──
# FIX del bug crítico de strip_tz anterior que operaba sobre rows/Series mixtas.
def _strip_tz_df(df: pd.DataFrame) -> pd.DataFrame:
    res = df.copy()
    for col in res.columns:
        if pd.api.types.is_datetime64_any_dtype(res[col]):
            if getattr(res[col].dt, 'tz', None) is not None:
                res[col] = res[col].dt.tz_convert(None)
    return res

# =====================================================================
# KPIs MATEMÁTICOS
# =====================================================================
def _merge_intervals(ivs: list) -> list:
    if not ivs: return []
    srt = sorted(ivs, key=lambda x: x[0])
    merged = [list(srt[0])]
    for s, e in srt[1:]:
        if s <= merged[-1][1]: merged[-1][1] = max(merged[-1][1], e)
        else: merged.append([s, e])
    return merged

def calc_kpis(df: pd.DataFrame, fecha_ini: date, fecha_fin: date) -> dict:
    h_tot = ((datetime.combine(fecha_fin, datetime_time(23,59,59)) -
              datetime.combine(fecha_ini, datetime_time(0,0,0))).total_seconds()) / 3600.0
    rng_s = SV_TZ.localize(datetime.combine(fecha_ini, datetime_time(0,0,0)))
    rng_e = SV_TZ.localize(datetime.combine(fecha_fin, datetime_time(23,59,59)))
    base = {
        "global": {"sla":100.0,"total_fallas":0,"mtbf":h_tot,"db":0.0,"mh":0.0,"p1":0},
        "t1":  {"fallas":0,"mttr":0.0,"acd":0.0,"clientes":0},
        "t2":  {"fallas":0,"mttr":0.0},
        "t3":  {"fallas":0,"clientes_est":0},
        "int": {"fallas":0,"mttr":0.0},
        "abiertos":0, "zonas_sla":{}
    }
    required = {'estado','inicio_incidente','categoria','duracion_horas',
                'clientes_afectados','conocimiento_tiempos','fin_incidente','causa_raiz','Severidad'}
    if df.empty or not required.issubset(df.columns): return base

    base["abiertos"] = int((df['estado'] == 'Abierto').sum())
    df_c = df[(df['estado'] == 'Cerrado') & df['inicio_incidente'].notnull()]
    df_int = df_c[df_c['categoria'] == CAT_INTERNA]
    df_ext = df_c[df_c['categoria'] != CAT_INTERNA]

    base["int"]["fallas"] = len(df_int)
    if not df_int.empty:
        iv = df_int[df_int['duracion_horas'] > 0]
        base["int"]["mttr"] = float(iv['duracion_horas'].mean()) if not iv.empty else 0.0
    if df_ext.empty: return base

    base["global"]["total_fallas"] = len(df_ext)
    base["global"]["p1"]           = int((df_ext['Severidad'] == '🔴 P1 (Crítica)').sum())
    df_t3 = df_ext[df_ext['conocimiento_tiempos'] != 'Total']
    base["t3"]["fallas"]       = len(df_t3)
    base["t3"]["clientes_est"] = int(df_t3['clientes_afectados'].sum())

    df_exact = df_ext[(df_ext['conocimiento_tiempos'] == 'Total') & df_ext['fin_incidente'].notnull()]
    if not df_exact.empty:
        df_sla = df_exact[df_exact['causa_raiz'] != 'Mantenimiento Programado']
        for z in df_sla['zona'].unique():
            df_z    = df_sla[df_sla['zona'] == z]
            s_cl    = df_z['inicio_incidente'].clip(lower=rng_s)
            e_cl    = df_z['fin_incidente'].clip(upper=rng_e)
            valid   = s_cl <= e_cl
            ivs_z   = [[s, e] for s, e in zip(s_cl[valid], e_cl[valid])]
            t_down_z = sum((e-s).total_seconds() for s,e in _merge_intervals(ivs_z)) / 3600.0
            base["zonas_sla"][z] = max(0.0, min(100.0, (h_tot - t_down_z) / h_tot * 100))
        base["global"]["db"] = float(df_exact['duracion_horas'].sum())
        base["global"]["mh"] = float(df_exact['duracion_horas'].max())

    df_t2 = df_exact[df_exact['clientes_afectados'] == 0]
    base["t2"]["fallas"] = len(df_t2)
    if not df_t2.empty: base["t2"]["mttr"] = float(df_t2['duracion_horas'].mean())

    df_t1 = df_exact[df_exact['clientes_afectados'] > 0]
    base["t1"]["fallas"] = len(df_t1)
    if not df_t1.empty:
        base["t1"]["mttr"]     = float(df_t1['duracion_horas'].mean())
        base["t1"]["clientes"] = int(df_t1['clientes_afectados'].sum())
        total_hc = (df_t1['duracion_horas'] * df_t1['clientes_afectados']).sum()
        base["t1"]["acd"] = float(total_hc / base["t1"]["clientes"]) if base["t1"]["clientes"] > 0 else 0.0

    if not df_exact.empty:
        df_sg = df_exact[df_exact['causa_raiz'] != 'Mantenimiento Programado']
        s_cl  = df_sg['inicio_incidente'].clip(lower=rng_s)
        e_cl  = df_sg['fin_incidente'].clip(upper=rng_e)
        valid = s_cl <= e_cl
        ivs   = [[s, e] for s, e in zip(s_cl[valid], e_cl[valid])]
        t_down = sum((e-s).total_seconds() for s,e in _merge_intervals(ivs)) / 3600.0
    else:
        t_down = 0.0
    base["global"]["sla"]  = max(0.0, min(100.0, (h_tot - t_down) / h_tot * 100)) if h_tot > 0 else 100.0
    base["global"]["mtbf"] = float((h_tot - t_down) / len(df_exact)) if len(df_exact) > 0 else float(h_tot)
    return base

# ── SLA histórico de los últimos 6 meses (cacheado) ──
@st.cache_data(ttl=300, show_spinner=False)
def get_sla_historico(m_actual: int, a_actual: int, cache_version: int = 0) -> pd.DataFrame:
    filas = []
    for i in range(5, -1, -1):
        m, a = m_actual - i, a_actual
        while m <= 0:
            m += 12; a -= 1
        fi = date(a, m, 1)
        ff = date(a, m, calendar.monthrange(a, m)[1])
        df_raw = _load_raw(fi, ff, False, "Todas", "Todos", "Todos")
        df_e   = _enriquecer(df_raw)
        k      = calc_kpis(df_e, fi, ff)
        filas.append({
            "Periodo": f"{MESES[m-1][:3]} {a}",
            "SLA (%)": round(k['global']['sla'], 3),
            "Fallas":  k['global']['total_fallas'],
            "MTTR (h)":round(k['t1']['mttr'], 2),
        })
    return pd.DataFrame(filas)

# =====================================================================
# GRÁFICOS
# =====================================================================
def dibujar_graficos(df_m: pd.DataFrame):
    if df_m.empty or 'duracion_horas' not in df_m.columns: return
    df_m = df_m.copy()

    st.markdown("### 🗺️ Análisis Geográfico y Temporal")
    zonas_coords = {z[0]: (z[1], z[2]) for z in get_zonas()}
    df_map = df_m.copy()
    df_map['lat'] = df_map['zona'].map(lambda x: zonas_coords.get(x, (13.6929, -89.2182))[0])
    df_map['lon'] = df_map['zona'].map(lambda x: zonas_coords.get(x, (13.6929, -89.2182))[1])
    agg = df_map.groupby(['zona_completa','lat','lon']).agg(
        Horas=('duracion_horas','sum'), Clientes=('clientes_afectados','sum')).reset_index()
    agg['Clientes_sz'] = agg['Clientes'].clip(lower=1)
    fig_map = px.scatter_mapbox(agg, lat="lat", lon="lon", hover_name="zona_completa",
                                size="Clientes_sz", color="Horas", color_continuous_scale="Inferno",
                                zoom=8.5, mapbox_style="carto-darkmatter",
                                labels={"Clientes_sz":"Clientes"}, title="Impacto Geográfico por Nodo")
    fig_map.update_layout(margin=dict(l=0,r=0,t=40,b=0), height=450)
    st.plotly_chart(fig_map, use_container_width=True)

    st.write("")
    dias_map   = {0:'Lunes',1:'Martes',2:'Miércoles',3:'Jueves',4:'Viernes',5:'Sábado',6:'Domingo'}
    dias_orden = list(dias_map.values())
    if 'inicio_incidente' in df_m.columns:
        df_heat = df_m[df_m['inicio_incidente'].notnull()].copy()
        if not df_heat.empty:
            df_heat['Día']  = pd.Categorical(df_heat['inicio_incidente'].dt.dayofweek.map(dias_map), categories=dias_orden, ordered=True)
            df_heat['Hora'] = df_heat['inicio_incidente'].dt.hour
            df_hm = df_heat.groupby(['Día','Hora'], observed=False).size().reset_index(name='Fallas')
            fig_hm = px.density_heatmap(df_hm, x='Hora', y='Día', z='Fallas', color_continuous_scale='Blues', nbinsx=24, title="Mapa de Calor (Concentración de Fallas por Hora)")
            fig_hm.update_layout(xaxis=dict(tickmode='linear',tick0=0,dtick=1), margin=dict(l=0,r=0,t=40,b=0), paper_bgcolor="rgba(0,0,0,0)", height=350)
            st.plotly_chart(fig_hm, use_container_width=True)

    st.write("")
    st.divider()
    st.markdown("### 📊 Responsabilidad y Causas Principales")
    c_pie, c_bar = st.columns(2)
    with c_pie:
        if 'es_externa' in df_m.columns:
            df_m['Tipo'] = df_m['es_externa'].map({True:'Externa (Fuerza Mayor)', False:'Interna (Infraestructura / NOC)'})
            agg_r = df_m.groupby('Tipo').size().reset_index(name='Eventos')
            fig_p = px.pie(agg_r, names='Tipo', values='Eventos', hole=0.5,
                           color_discrete_sequence=[COLOR_TEAL, COLOR_DANGER], title="Tasa de Responsabilidad")
            fig_p.update_traces(textinfo='percent', textposition='inside')
            fig_p.update_layout(showlegend=True, legend=dict(orientation="h",yanchor="top",y=-0.1,xanchor="center",x=0.5),
                                 margin=dict(l=0,r=0,t=40,b=0), paper_bgcolor="rgba(0,0,0,0)", height=400)
            st.plotly_chart(fig_p, use_container_width=True)
    with c_bar:
        if 'causa_raiz' in df_m.columns:
            dc = df_m.groupby('causa_raiz').size().reset_index(name='Alertas').sort_values('Alertas', ascending=True).tail(8)
            fig_b = px.bar(dc, x='Alertas', y='causa_raiz', orientation='h',
                           color_discrete_sequence=[COLOR_PRIMARY], text_auto='.0f', title="Top Causas Raíz")
            fig_b.update_traces(textposition='outside')
            fig_b.update_layout(margin=dict(l=0,r=0,t=40,b=0), paper_bgcolor="rgba(0,0,0,0)", height=400, xaxis_title="", yaxis_title="")
            st.plotly_chart(fig_b, use_container_width=True)

# =====================================================================
# EXCEL EXPORT CON FORMATO
# Hojas: Resumen, T1, T2, T3, Internos — con colores por severidad
# =====================================================================
def generar_excel(df: pd.DataFrame, kpis: dict, label_periodo: str) -> bytes:
    buf = BytesIO()
    drop_cols = ['Severidad','es_externa','zona_completa','deleted_at','impacto_porcentaje']

    COLOR_SEV = {
        '🔴 P1 (Crítica)':    'FFFF2B2B',
        '🟠 P2 (Alta)':       'FFFF9F43',
        '🟡 P3 (Media)':      'FFFFFF99',
        '🟢 P4 (Interna)':    'FFD5F5E3',
        '🚨 CRÍTICA (En Curso)': 'FFFF6B6B',
    }
    HDR_FILL  = 'FF1D2C59'
    HDR_FONT  = 'FFFFFFFF'

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        thin = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0'),
        )

        def _add_sheet(ws, df_s: pd.DataFrame, title: str):
            ws.title = title
            # Cabecera corporativa
            ws['A1'] = "MULTINET — Network Operations Center"
            ws['A1'].font = Font(bold=True, size=13, color=HDR_FONT)
            ws['A1'].fill = PatternFill("solid", fgColor=HDR_FILL)
            ws.merge_cells(f'A1:{get_column_letter(max(1, len(df_s.columns)))}1')
            ws['A2'] = f"Periodo: {label_periodo}   |   Hoja: {title}"
            ws['A2'].font = Font(italic=True, color='FF888888', size=9)
            ws.merge_cells(f'A2:{get_column_letter(max(1, len(df_s.columns)))}2')

            if df_s.empty:
                ws['A3'] = "Sin registros en este nivel."
                return

            cols = [c for c in df_s.columns if c not in drop_cols]
            sev_col = 'Severidad'

            # Encabezados
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=3, column=ci, value=str(col))
                cell.font      = Font(bold=True, color=HDR_FONT, size=9)
                cell.fill      = PatternFill("solid", fgColor=HDR_FILL)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = thin

            # Datos con colores de severidad
            sev_data = df_s[sev_col].values if sev_col in df_s.columns else None
            for ri, row_data in enumerate(df_s[cols].itertuples(index=False), 4):
                sev_val  = sev_data[ri-4] if sev_data is not None else None
                fill_hex = COLOR_SEV.get(sev_val, 'FFFFFFFF') if sev_val else 'FFFFFFFF'
                bg_fill  = PatternFill("solid", fgColor=fill_hex)
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=ri, column=ci,
                                   value=str(val) if isinstance(val, (list, dict)) else val)
                    cell.fill      = bg_fill
                    cell.border    = thin
                    cell.font      = Font(size=8)
                    cell.alignment = Alignment(vertical='center')

            # Auto-ancho de columnas
            for ci, col in enumerate(cols, 1):
                max_len = max(len(str(col)), *(len(str(ws.cell(r, ci).value or '')) for r in range(3, ws.max_row+1)))
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)
            ws.row_dimensions[3].height = 18

        # Filtrar niveles
        def _filt(df_in, cond):
            return df_in[cond].drop(columns=drop_cols, errors='ignore') if not df_in.empty else pd.DataFrame()

        if df.empty or 'conocimiento_tiempos' not in df.columns:
            df_t1 = df_t2 = df_t3 = df_int = pd.DataFrame()
        else:
            df_ext_e = df[(df.get('estado', pd.Series()) == 'Cerrado') & (df.get('categoria', pd.Series()) != CAT_INTERNA)] if 'estado' in df.columns and 'categoria' in df.columns else pd.DataFrame()
            df_int_e = df[df.get('categoria', pd.Series()) == CAT_INTERNA] if 'categoria' in df.columns else pd.DataFrame()
            df_exact = df_ext_e[(df_ext_e['conocimiento_tiempos'] == 'Total')] if not df_ext_e.empty else pd.DataFrame()
            df_t1    = df_exact[df_exact['clientes_afectados'] > 0].drop(columns=drop_cols, errors='ignore') if not df_exact.empty else pd.DataFrame()
            df_t2    = df_exact[df_exact['clientes_afectados'] == 0].drop(columns=drop_cols, errors='ignore') if not df_exact.empty else pd.DataFrame()
            df_t3    = df_ext_e[df_ext_e['conocimiento_tiempos'] != 'Total'].drop(columns=drop_cols, errors='ignore') if not df_ext_e.empty else pd.DataFrame()
            df_int   = df_int_e.drop(columns=drop_cols, errors='ignore') if not df_int_e.empty else pd.DataFrame()

        # Hoja Resumen KPIs
        ws_res = wb.active
        ws_res.title = "Resumen KPIs"
        ws_res['A1'] = "MULTINET — Resumen Ejecutivo"
        ws_res['A1'].font = Font(bold=True, size=14, color=HDR_FONT)
        ws_res['A1'].fill = PatternFill("solid", fgColor=HDR_FILL)
        ws_res.merge_cells('A1:C1')
        ws_res['A2'] = f"Periodo: {label_periodo}"
        ws_res['A2'].font = Font(italic=True, color='FF888888', size=9)
        ws_res.merge_cells('A2:C2')
        encabezados_kpi = [("Indicador", "Valor", "Descripción")]
        filas_kpi = [
            ("SLA Global",         f"{kpis['global']['sla']:.4f}%",    "Excluye mantenimientos"),
            ("MTTR Real (T1)",     f"{kpis['t1']['mttr']:.2f} h",      "Promedio resolución con clientes"),
            ("ACD Afectación",     f"{kpis['t1']['acd']:.2f} h",       "Percepción real del cliente"),
            ("Impacto Acumulado",  f"{kpis['global']['db']/24:.3f} días","Total horas / 24"),
            ("Falla Mayor",        f"{kpis['global']['mh']:.2f} h",    "Incidente más largo"),
            ("MTBF",               f"{kpis['global']['mtbf']:.1f} h",  "Tiempo medio entre fallas"),
            ("Fallas P1",          str(kpis['global']['p1']),           "≥12h o ≥1000 clientes"),
            ("Total Eventos",      str(kpis['global']['total_fallas']), "Externos cerrados"),
            ("Clientes Afect.",    f"{kpis['t1']['clientes']:,}",       "T1 exactos"),
            ("Tickets Abiertos",   str(kpis['abiertos']),              "Pendientes de cierre"),
        ]
        for ci, h in enumerate(encabezados_kpi[0], 1):
            c = ws_res.cell(row=3, column=ci, value=h)
            c.font = Font(bold=True, color=HDR_FONT, size=9)
            c.fill = PatternFill("solid", fgColor=HDR_FILL)
            c.alignment = Alignment(horizontal='center')
            c.border = thin
        for ri, (ind, val, desc) in enumerate(filas_kpi, 4):
            for ci, v in enumerate([ind, val, desc], 1):
                c = ws_res.cell(row=ri, column=ci, value=v)
                c.fill   = PatternFill("solid", fgColor='FFF5F5F5' if ri % 2 == 0 else 'FFFFFFFF')
                c.border = thin; c.font = Font(size=8)
        ws_res.column_dimensions['A'].width = 22
        ws_res.column_dimensions['B'].width = 16
        ws_res.column_dimensions['C'].width = 35

        _add_sheet(wb.create_sheet("T1 - Con Clientes"),  df_t1, "T1 - Con Clientes")
        _add_sheet(wb.create_sheet("T2 - Sin Clientes"),  df_t2, "T2 - Sin Clientes")
        _add_sheet(wb.create_sheet("T3 - Parciales"),     df_t3, "T3 - Parciales")
        _add_sheet(wb.create_sheet("Internos"),           df_int, "Internos")

        wb.save(buf)
    except ImportError:
        # Fallback a CSV si openpyxl no está disponible
        df.drop(columns=drop_cols, errors='ignore').to_csv(buf, index=False)

    buf.seek(0)
    return buf.getvalue()

# =====================================================================
# AUDITORÍA
# =====================================================================
def log_audit(action: str, detail: str):
    try:
        with engine.begin() as conn:
            conn.execute(
                text("INSERT INTO audit_logs (timestamp,username,action,details) VALUES (:t,:u,:a,:d)"),
                {"t": datetime.now(SV_TZ).replace(tzinfo=None),
                 "u": st.session_state.get("username","?"), "a": action, "d": detail}
            )
    except Exception: pass

# =====================================================================
# PDF EJECUTIVO
# FIX #4: usa df['categoria'] con verificación de columna (no df.get())
# FIX #6: h_tot_periodo usa fechas reales del periodo, no date.today()
# =====================================================================
def generar_pdf(label_periodo: str, kpis: dict, df: pd.DataFrame,
                zonas_todas: list, fecha_ini: date, fecha_fin: date) -> bytes:
    def mk(name, size, font='Helvetica', color=rl_colors.black, **kw):
        return ParagraphStyle(name, fontSize=size, fontName=font, textColor=color, **kw)

    COL_H   = rl_colors.HexColor(COLOR_SECONDARY)
    COL_H2  = rl_colors.HexColor(COLOR_TEAL)
    COL_ACC = rl_colors.HexColor(COLOR_PRIMARY)
    COL_GRY = rl_colors.HexColor('#888888')
    COL_LGT = rl_colors.HexColor('#f5f5f5')
    COL_WH  = rl_colors.white
    COL_BLK = rl_colors.black

    s_brand = mk('br', 26, 'Helvetica-Bold', rl_colors.HexColor(COLOR_SECONDARY), spaceAfter=0)
    s_title = mk('t',  13, 'Helvetica-Bold', rl_colors.HexColor(COLOR_PRIMARY),   spaceAfter=2)
    s_body  = mk('b',   9, 'Helvetica',      COL_BLK, spaceAfter=3)
    s_sec   = mk('h',  12, 'Helvetica-Bold', COL_H,   spaceBefore=12, spaceAfter=5)
    s_foot  = mk('f',   7, 'Helvetica',      COL_GRY)
    s_period= mk('pe', 11, 'Helvetica-Bold', COL_ACC, spaceAfter=2)

    def tbl_style(hdr_color, accent=False):
        base = [
            ('BACKGROUND',    (0,0),(-1,0),  hdr_color),
            ('TEXTCOLOR',     (0,0),(-1,0),  COL_WH),
            ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),(-1,0),  8),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [COL_LGT, COL_WH]),
            ('FONTNAME',      (0,1),(-1,-1), 'Helvetica'),
            ('FONTSIZE',      (0,1),(-1,-1), 8),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('GRID',          (0,0),(-1,-1), 0.3, rl_colors.HexColor('#dddddd')),
            ('ROWHEIGHT',     (0,0),(-1,-1), 18),
            ('LEFTPADDING',   (0,0),(-1,-1), 6),
            ('RIGHTPADDING',  (0,0),(-1,-1), 6),
        ]
        if accent:
            base += [('TEXTCOLOR',(1,1),(1,-1),rl_colors.HexColor(COLOR_PRIMARY)),
                     ('FONTNAME', (1,1),(1,-1),'Helvetica-Bold'),
                     ('ALIGN',    (1,0),(1,-1),'CENTER')]
        return TableStyle(base)

    def sla_color(v):
        if v >= 99: return rl_colors.HexColor('#29b09d')
        if v >= 95: return rl_colors.HexColor('#ff9f43')
        return rl_colors.HexColor(COLOR_DANGER)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)
    story = []
    now_str = datetime.now(SV_TZ).strftime('%d/%m/%Y %H:%M')

    hdr_data = [[Paragraph("MULTINET", s_brand),
                 Paragraph(f"<b>Periodo:</b> {label_periodo}<br/>"
                           f"<b>Generado:</b> {now_str} (hora El Salvador)<br/>"
                           f"<b>Clasificación:</b> Documento Confidencial", s_body)]]
    hdr_t = Table(hdr_data, colWidths=[7*cm, 10*cm])
    hdr_t.setStyle(TableStyle([
        ('VALIGN',   (0,0),(-1,-1),'MIDDLE'), ('ALIGN',(1,0),(1,0),'RIGHT'),
        ('LINEBELOW',(0,0),(-1,0), 1.5, rl_colors.HexColor(COLOR_PRIMARY)),
        ('BOTTOMPADDING',(0,0),(-1,0),8),
    ]))
    story.append(hdr_t)
    story.append(Paragraph("Reporte Ejecutivo — Network Operations Center", s_title))
    story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor('#dddddd'), spaceAfter=10))

    sla_val = kpis['global']['sla']
    story.append(Paragraph("1. Indicadores Operativos Globales", s_sec))
    kpi_rows = [
        ['Indicador', 'Valor', 'Referencia / Descripción'],
        ['SLA Global (Disponibilidad)',  f"{sla_val:.4f}%",                'Meta: ≥ 99.5 % | Excluye mantenimientos programados'],
        ['MTTR Real (T1)',               f"{kpis['t1']['mttr']:.2f} h",   'Resolución promedio con clientes afectados'],
        ['ACD — Afectación cliente',     f"{kpis['t1']['acd']:.2f} h",    'Percepción real: (Σ dur×clientes) / total clientes'],
        ['Impacto Acumulado',            f"{kpis['global']['db']/24:.3f} días", 'Total horas de caída / 24'],
        ['Falla Mayor (Pico)',           f"{kpis['global']['mh']:.2f} h", 'Incidente individual más largo'],
        ['MTBF (Estabilidad de Red)',    f"{kpis['global']['mtbf']:.1f} h",'Horas disponibles / cantidad de fallas'],
        ['Fallas P1 Críticas',          str(kpis['global']['p1']),        'Duración ≥12h  O  clientes ≥1000'],
        ['Total Eventos Externos',      str(kpis['global']['total_fallas']),'Incidentes cerrados con impacto'],
        ['Clientes Impactados (T1)',     f"{kpis['t1']['clientes']:,}",    'Usuarios en incidentes T1 exactos'],
    ]
    t_kpi = Table(kpi_rows, colWidths=[6*cm, 3*cm, 8.5*cm])
    t_kpi.setStyle(tbl_style(COL_H, accent=True))
    t_kpi.setStyle(TableStyle([('BACKGROUND',(1,1),(1,1),sla_color(sla_val)),('TEXTCOLOR',(1,1),(1,1),COL_WH)]))
    story += [t_kpi, Spacer(1, 0.3*cm)]

    story.append(Paragraph("2. Disponibilidad por Zona / Nodo", s_sec))
    zona_rows = [['Zona / Nodo', 'SLA (%)', 'Estado']]
    for z_nombre, _, _ in zonas_todas:
        sla_z = kpis['zonas_sla'].get(z_nombre, 100.0)
        estado_z = "Excelente" if sla_z >= 99.0 else ("Aceptable" if sla_z >= 95.0 else "Crítico")
        zona_rows.append([z_nombre, f"{sla_z:.3f}%", estado_z])
    t_zona = Table(zona_rows, colWidths=[8*cm, 3.5*cm, 4*cm])
    t_zona.setStyle(tbl_style(COL_H2))
    for ri in range(1, len(zona_rows)):
        v = float(zona_rows[ri][1].replace('%',''))
        c = sla_color(v)
        t_zona.setStyle(TableStyle([('TEXTCOLOR',(1,ri),(1,ri),c),('FONTNAME',(1,ri),(1,ri),'Helvetica-Bold'),
                                     ('TEXTCOLOR',(2,ri),(2,ri),c),('FONTNAME',(2,ri),(2,ri),'Helvetica-Bold')]))
    story += [t_zona, Spacer(1, 0.3*cm)]

    story.append(Paragraph("3. Estratificación por Nivel de Conocimiento", s_sec))
    tier_rows = [['Nivel','Eventos','MTTR','Clientes','Notas'],
                 ['T1 (Exactos + Clientes)',   str(kpis['t1']['fallas']),  f"{kpis['t1']['mttr']:.2f}h", f"{kpis['t1']['clientes']:,}",'Base principal SLA/ACD'],
                 ['T2 (Exactos sin Clientes)', str(kpis['t2']['fallas']),  f"{kpis['t2']['mttr']:.2f}h", '0','Infraestructura sin impacto directo'],
                 ['T3 (Parciales)',            str(kpis['t3']['fallas']),  '—',                           f"~{kpis['t3']['clientes_est']:,}",'Excluidos del SLA'],
                 ['T-Int (Internos)',          str(kpis['int']['fallas']), f"{kpis['int']['mttr']:.2f}h", '0','Fallas internas NOC']]
    t_tier = Table(tier_rows, colWidths=[5*cm, 2*cm, 2.5*cm, 2.5*cm, 5.5*cm])
    t_tier.setStyle(tbl_style(COL_H))
    story += [t_tier, Spacer(1, 0.3*cm)]

    story.append(Paragraph("4. Salud de Datos y Alertas", s_sec))
    pend_rows = [['Indicador','Cantidad','Acción recomendada'],
                 ['Tickets Abiertos',               str(kpis['abiertos']),       'Cerrar con hora exacta'],
                 ['Sin tiempos exactos (T3)',        str(kpis['t3']['fallas']),  'Completar para incluir en SLA'],
                 ['Incidentes internos registrados', str(kpis['int']['fallas']), 'Revisar causa raíz']]
    t_pend = Table(pend_rows, colWidths=[6*cm, 2.5*cm, 9*cm])
    t_pend.setStyle(tbl_style(COL_H2))
    story += [t_pend, Spacer(1, 0.3*cm)]

    sec_num = 5
    if not df.empty and 'duracion_horas' in df.columns and 'estado' in df.columns and 'conocimiento_tiempos' in df.columns:
        df_top = df[(df['estado']=='Cerrado') & (df['conocimiento_tiempos']=='Total')]
        if 'categoria' in df_top.columns:
            df_top = df_top[df_top['categoria'] != CAT_INTERNA]
        if not df_top.empty:
            story.append(Paragraph(f"{sec_num}. Top 8 Incidentes por Duración", s_sec))
            top8 = df_top.nlargest(8, 'duracion_horas')
            inc_rows = [['ID','Zona','Causa Raíz','Dur.','Clientes','Sev.']]
            for _, r in top8.iterrows():
                sev_text = re.sub(r'[^\x20-\x7E]','', str(r.get('Severidad','')))[:12]
                inc_rows.append([str(int(r.get('id',0))), str(r.get('zona',''))[:20],
                                 str(r.get('causa_raiz',''))[:28], f"{r.get('duracion_horas',0):.1f}h",
                                 str(int(r.get('clientes_afectados',0))), sev_text])
            t_inc = Table(inc_rows, colWidths=[1.2*cm,4*cm,5*cm,2*cm,2*cm,3.3*cm])
            t_inc.setStyle(tbl_style(COL_H))
            story += [t_inc, Spacer(1, 0.3*cm)]
            sec_num += 1

    if not df.empty and 'causa_raiz' in df.columns and 'estado' in df.columns:
        df_causas = df[df['estado']=='Cerrado']
        if not df_causas.empty:
            story.append(Paragraph(f"{sec_num}. Análisis de Causas Raíz", s_sec))
            causas_map = get_causas_con_flag()
            top_c = df_causas.groupby('causa_raiz').agg(Eventos=('id','count'),Horas=('duracion_horas','sum'),Clientes=('clientes_afectados','sum')).reset_index().sort_values('Eventos',ascending=False).head(8)
            caus_rows = [['Causa Raíz','Eventos','Horas','Clientes','Tipo']]
            for _, r in top_c.iterrows():
                caus_rows.append([str(r['causa_raiz'])[:35], str(int(r['Eventos'])),
                                  f"{r['Horas']:.1f}h", str(int(r['Clientes'])),
                                  "Externa" if causas_map.get(r['causa_raiz'],False) else "Interna"])
            t_caus = Table(caus_rows, colWidths=[6.5*cm,2*cm,2.5*cm,2.5*cm,2.5*cm])
            t_caus.setStyle(tbl_style(COL_H))
            story += [t_caus, Spacer(1, 0.3*cm)]
            sec_num += 1

    # FIX #4 + #6: verificación de columna y horas reales del periodo
    if not df.empty and 'duracion_horas' in df.columns and 'zona_completa' in df.columns and 'estado' in df.columns:
        df_ext_pdf = pd.DataFrame()
        if 'categoria' in df.columns:
            df_ext_pdf = df[(df['categoria'] != CAT_INTERNA) & (df['estado'] == 'Cerrado')]
        if not df_ext_pdf.empty:
            story.append(Paragraph(f"{sec_num}. Zonas con Mayor Afectación", s_sec))
            # FIX #6: usar horas reales del periodo, no date.today()
            h_periodo_real = ((datetime.combine(fecha_fin, datetime_time(23,59,59)) -
                               datetime.combine(fecha_ini, datetime_time(0,0,0))).total_seconds()) / 3600.0
            top_z = df_ext_pdf.groupby('zona_completa')['duracion_horas'].sum().nlargest(8).reset_index()
            z_rows = [['Zona / Nodo','Horas','Días Equiv.','% Periodo']]
            for _, r in top_z.iterrows():
                pct = min(100.0, r['duracion_horas'] / h_periodo_real * 100) if h_periodo_real > 0 else 0.0
                z_rows.append([str(r['zona_completa']), f"{r['duracion_horas']:.1f}h",
                                f"{r['duracion_horas']/24:.2f}", f"{pct:.2f}%"])
            t_z = Table(z_rows, colWidths=[6.5*cm,2.5*cm,3*cm,3.5*cm])
            t_z.setStyle(tbl_style(COL_H2, accent=True))
            story += [t_z, Spacer(1, 0.3*cm)]

    story += [
        Spacer(1, 0.8*cm),
        HRFlowable(width="100%", thickness=0.5, color=rl_colors.HexColor('#dddddd'), spaceAfter=5),
        Paragraph(f"MULTINET NOC  ·  {now_str}  ·  Periodo: {label_periodo}  ·  Documento Confidencial — Uso Interno", s_foot),
    ]
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

# FIX #7: PDF cacheado en session_state — se regenera solo cuando cambian datos o periodo.
def _get_pdf_cached(label_periodo, kpis, df_m, zonas, fecha_ini, fecha_fin):
    cache_key = f"{label_periodo}_{_dv()}"
    if st.session_state.pdf_cache_key != cache_key:
        st.session_state.pdf_cache_bytes = generar_pdf(label_periodo, kpis, df_m, zonas, fecha_ini, fecha_fin)
        st.session_state.pdf_cache_key   = cache_key
    return st.session_state.pdf_cache_bytes

# =====================================================================
# LOGIN SEGURO
# FIX #5: _invalidar_cache() tras login exitoso para evitar datos viejos
# =====================================================================
def do_login():
    u, p = st.session_state.log_u, st.session_state.log_p
    try:
        with engine.begin() as conn:
            ud = conn.execute(
                text("SELECT id,password_hash,role,failed_attempts,locked_until,is_banned FROM users WHERE username=:u"),
                {"u": u}
            ).fetchone()
            if ud:
                uid, ph, rol, fa, ldt, ban = ud
                fa = fa or 0
                now_sv = datetime.now(SV_TZ).replace(tzinfo=None)
                if ban:
                    st.session_state.log_err = "❌ Cuenta baneada permanentemente."
                elif ldt and ldt > now_sv:
                    st.session_state.log_err = f"⏳ Bloqueado por {(ldt - now_sv).seconds // 60 + 1} min."
                elif check_pw(p, ph):
                    conn.execute(text("UPDATE users SET failed_attempts=0,locked_until=NULL WHERE id=:id"), {"id": uid})
                    st.session_state.update({"logged_in": True, "role": rol, "username": u, "log_err": ""})
                    _invalidar_cache()  # FIX #5
                    return
                else:
                    fa += 1
                    if fa >= 6:
                        conn.execute(text("UPDATE users SET is_banned=TRUE,failed_attempts=:f WHERE id=:id"), {"f": fa,"id": uid})
                        st.session_state.log_err = "❌ Cuenta bloqueada permanentemente por seguridad."
                    elif fa % 3 == 0:
                        conn.execute(text("UPDATE users SET locked_until=:dt,failed_attempts=:f WHERE id=:id"),
                                     {"dt": now_sv + timedelta(minutes=5),"f": fa,"id": uid})
                        st.session_state.log_err = "⏳ Demasiados intentos. Bloqueado por 5 min."
                    else:
                        conn.execute(text("UPDATE users SET failed_attempts=:f WHERE id=:id"), {"f": fa,"id": uid})
                        st.session_state.log_err = "❌ Credenciales incorrectas."
            else:
                st.session_state.log_err = "❌ Credenciales incorrectas."
    except Exception:
        st.session_state.log_err = "Error de conexión DB."
    st.session_state.log_u = ""; st.session_state.log_p = ""

if not st.session_state.logged_in:
    st.markdown("<div style='margin-top:15vh;'></div>", unsafe_allow_html=True)
    _, col_c, _ = st.columns([1, 1.2, 1])
    with col_c:
        with st.container(border=True):
            st.markdown("<div style='text-align:center;padding:20px 0 10px 0;'><div style='font-size:46px;'>🔐</div><h2 style='margin:10px 0 4px;color:#fff;font-weight:700;'>Acceso NOC Central</h2></div>", unsafe_allow_html=True)
            if st.session_state.log_err:
                st.error(st.session_state.log_err, icon="⚠️")
                st.session_state.log_err = ""
            st.text_input("Usuario",    key="log_u")
            st.text_input("Contraseña", key="log_p", type="password")
            st.button("Iniciar Sesión", type="primary", on_click=do_login, use_container_width=True)
            st.caption("Contacte al Administrador de Redes para gestionar sus credenciales.")
    st.stop()

# =====================================================================
# SIDEBAR
# =====================================================================
with st.sidebar:
    st.caption(f"👤 **{st.session_state.username}** ({st.session_state.role.capitalize()})  |  NOC v1.0")
    st.divider()

    anio_act   = datetime.now(SV_TZ).year
    anios_list = sorted({anio_act+1, anio_act, anio_act-1, anio_act-2}, reverse=True)
    idx_anio   = anios_list.index(anio_act) if anio_act in anios_list else 0

    a_sel = st.selectbox("🗓️ Año",  anios_list, index=idx_anio)
    m_sel = st.selectbox("📅 Mes",  MESES, index=datetime.now(SV_TZ).month - 1)
    m_idx = MESES.index(m_sel) + 1

    fecha_ini     = date(a_sel, m_idx, 1)
    fecha_fin     = date(a_sel, m_idx, calendar.monthrange(a_sel, m_idx)[1])
    label_periodo = f"{m_sel} {a_sel}"

    st.divider()
    z_sel   = st.selectbox("🗺️ Zona",     ["Todas"]  + [z[0] for z in get_zonas()])
    srv_sel = st.selectbox("🌐 Servicio", ["Todos"]  + get_cat("cat_servicios"))
    seg_sel = st.selectbox("🏢 Segmento", ["Todos"]  + list(DEFAULT_CATEGORIAS))

    df_m = load_and_enrich(fecha_ini, fecha_fin, False, z_sel, srv_sel, seg_sel, cache_version=_dv())

    st.divider()
    if not df_m.empty:
        kpis_sidebar = calc_kpis(df_m, fecha_ini, fecha_fin)
        # FIX #7: PDF generado lazy y cacheado en session_state
        pdf_bytes = _get_pdf_cached(label_periodo, kpis_sidebar, df_m, get_zonas(), fecha_ini, fecha_fin)
        st.download_button(
            "📥 Descargar Reporte PDF",
            data=pdf_bytes,
            file_name=f"Reporte_NOC_{m_sel}_{a_sel}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

    st.divider()
    if st.button("🚪 Cerrar Sesión", use_container_width=True):
        log_audit("LOGOUT", "Sesión cerrada.")
        st.session_state.clear()
        st.rerun()

    st.markdown("""
        <div style="margin-top:50px;text-align:center;color:#666;font-size:11px;">
            💻 Engineered by<br><b>Luis Salvador Guzmán López</b>
        </div>
    """, unsafe_allow_html=True)

# =====================================================================
# PESTAÑAS POR ROL
# =====================================================================
role  = st.session_state.role
t_idx = 0

if   role == 'admin':   tab_labels = ["📊 Dashboard", "📝 Registrar Evento", "🗂️ Historial y Edición", "⚙️ Configuración"]
elif role == 'auditor': tab_labels = ["📊 Dashboard", "📝 Registrar Evento", "🗂️ Historial y Edición"]
else:                   tab_labels = ["📊 Dashboard"]

tabs = st.tabs(tab_labels)

# ─────────────────────────────────────────────
# TAB 0 — DASHBOARD
# ─────────────────────────────────────────────
with tabs[t_idx]:
    st.title(f"📊 Rendimiento de Red: {label_periodo}")

    if df_m.empty:
        st.success("🟢 Sin incidentes registrados en los filtros seleccionados. ¡La red está al 100%!")
    else:
        kpis = calc_kpis(df_m, fecha_ini, fecha_fin)

        # ── NUEVA FEATURE: 🔔 Badge de alerta en título del navegador ──
        n_abiertos = kpis.get('abiertos', 0)
        if n_abiertos > 0:
            st.markdown(
                f"<script>document.title = '🚨 ({n_abiertos}) Multinet NOC';</script>",
                unsafe_allow_html=True
            )

        # Delta vs mes anterior
        p_m_idx = 12 if m_idx == 1 else m_idx - 1
        p_a_sel = a_sel - 1 if m_idx == 1 else a_sel
        p_fi    = date(p_a_sel, p_m_idx, 1)
        p_ff    = date(p_a_sel, p_m_idx, calendar.monthrange(p_a_sel, p_m_idx)[1])
        df_prev = load_and_enrich(p_fi, p_ff, False, z_sel, srv_sel, seg_sel, cache_version=_dv())
        kpis_prev = calc_kpis(df_prev, p_fi, p_ff) if not df_prev.empty else None

        def _delta(cat, key, divisor=1, fmt="{:+.2f}", suffix=""):
            if not kpis_prev: return None
            return f"{fmt.format((kpis[cat][key] - kpis_prev[cat][key]) / divisor)}{' '+suffix if suffix else ''}"

        if 'estado' in df_m.columns:
            df_abiertos = df_m[df_m['estado'] == 'Abierto']
            if not df_abiertos.empty:
                st.error(f"🚨 Tienes {len(df_abiertos)} Falla(s) en Curso (Tickets Abiertos).", icon="🚨")
                cols_show = [c for c in ['zona_completa','equipo_afectado','inicio_incidente','causa_raiz','descripcion'] if c in df_abiertos.columns]
                st.dataframe(df_abiertos[cols_show].rename(columns={
                    'zona_completa':'Nodo','equipo_afectado':'Equipo',
                    'inicio_incidente':'Hora de Caída','causa_raiz':'Diagnóstico'}),
                    hide_index=True, use_container_width=True)
                st.divider()

        st.markdown("### 📈 Estado General de la Red")
        st.caption("*Métricas globales calculadas excluyendo los Mantenimientos Programados.*")

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("SLA Global",        f"{kpis['global']['sla']:.3f}%",
                  delta=_delta('global','sla', fmt="{:+.3f}", suffix="%"), delta_color="normal")
        c2.metric("Total Eventos",      kpis['global']['total_fallas'],
                  delta=_delta('global','total_fallas', fmt="{:+.0f}"), delta_color="inverse")
        c3.metric("MTBF (Estabilidad)", f"{kpis['global']['mtbf']:.1f} horas",
                  delta=_delta('global','mtbf', fmt="{:+.1f}", suffix="h"), delta_color="normal")
        c4.metric("Impacto Acumulado",  f"{kpis['global']['db']/24:.2f} días",
                  delta=_delta('global','db', divisor=24, fmt="{:+.2f}", suffix="d"), delta_color="inverse")
        c5.metric("Falla Mayor (Pico)", f"{kpis['global']['mh']:.1f} horas",
                  delta=_delta('global','mh', fmt="{:+.1f}", suffix="h"), delta_color="inverse")

        st.divider()

        # ── NUEVA FEATURE: 📈 Tendencia SLA histórica ──
        st.markdown("### 📈 Tendencia de SLA — Últimos 6 Meses")
        st.caption("*Seguimiento automático de disponibilidad para detectar tendencias de mejora o deterioro.*")
        df_hist = get_sla_historico(m_idx, a_sel, cache_version=_dv())
        if not df_hist.empty:
            fig_trend = go.Figure()
            fig_trend.add_trace(go.Scatter(
                x=df_hist["Periodo"], y=df_hist["SLA (%)"],
                mode='lines+markers+text', name='SLA (%)',
                line=dict(color=COLOR_PRIMARY, width=3),
                marker=dict(size=9, color=COLOR_PRIMARY),
                text=[f"{v:.2f}%" for v in df_hist["SLA (%)"]],
                textposition='top center', textfont=dict(size=10),
            ))
            fig_trend.add_hline(y=99.5, line_dash="dot", line_color=COLOR_TEAL, annotation_text="Meta 99.5%", annotation_position="left")
            fig_trend.add_hline(y=99.0, line_dash="dash", line_color="#ff9f43", annotation_text="99.0%", annotation_position="left")
            fig_trend.update_layout(
                yaxis=dict(range=[max(0, df_hist["SLA (%)"].min() - 2), 100.2], title="SLA (%)"),
                xaxis_title="Periodo", margin=dict(l=0,r=0,t=30,b=0),
                paper_bgcolor="rgba(0,0,0,0)", height=280,
                legend=dict(orientation="h", y=1.1)
            )
            st.plotly_chart(fig_trend, use_container_width=True)

            # Mini KPIs de la tendencia
            tc1, tc2, tc3, tc4 = st.columns(4)
            sla_prom  = df_hist["SLA (%)"].mean()
            sla_trend = df_hist["SLA (%)"].iloc[-1] - df_hist["SLA (%)"].iloc[0] if len(df_hist) > 1 else 0
            tc1.metric("SLA Promedio 6M", f"{sla_prom:.3f}%")
            tc2.metric("Tendencia",        f"{sla_trend:+.3f}%", delta_color="normal" if sla_trend >= 0 else "inverse")
            tc3.metric("Mejor Mes",        df_hist.loc[df_hist['SLA (%)'].idxmax(), 'Periodo'])
            tc4.metric("Peor Mes",         df_hist.loc[df_hist['SLA (%)'].idxmin(), 'Periodo'])

        st.divider()

        # ── NUEVA FEATURE: 🗺️ Modo Comparativo de dos meses ──
        comparar = st.toggle("🗺️ Activar Modo Comparativo (comparar con otro periodo)")
        if comparar:
            st.markdown("#### 📊 Comparativo de Periodos")
            cc1, cc2 = st.columns(2)
            with cc1:
                comp_m_sel = st.selectbox("📅 Mes de comparación", MESES,
                                           index=(m_idx - 2) % 12, key="comp_mes")
                comp_a_sel = st.selectbox("🗓️ Año de comparación", anios_list,
                                           index=idx_anio, key="comp_anio")
            comp_m_idx = MESES.index(comp_m_sel) + 1
            comp_fi    = date(comp_a_sel, comp_m_idx, 1)
            comp_ff    = date(comp_a_sel, comp_m_idx, calendar.monthrange(comp_a_sel, comp_m_idx)[1])
            df_comp    = load_and_enrich(comp_fi, comp_ff, False, z_sel, srv_sel, seg_sel, cache_version=_dv())
            kpis_comp  = calc_kpis(df_comp, comp_fi, comp_ff)

            col_a, col_sep, col_b = st.columns([5, 0.2, 5])
            with col_a:
                st.markdown(f"**{label_periodo}** (actual)")
                st.metric("SLA",             f"{kpis['global']['sla']:.3f}%")
                st.metric("MTTR T1",         f"{kpis['t1']['mttr']:.2f} h")
                st.metric("ACD",             f"{kpis['t1']['acd']:.2f} h")
                st.metric("Eventos",         kpis['global']['total_fallas'])
                st.metric("Clientes Impact.", f"{kpis['t1']['clientes']:,}")
                st.metric("Impacto Acum.",   f"{kpis['global']['db']/24:.2f} días")
                st.metric("MTBF",            f"{kpis['global']['mtbf']:.1f} h")
                st.metric("Fallas P1",       kpis['global']['p1'])
            col_sep.markdown("<div style='border-left:2px solid #333;height:350px;margin:auto;'></div>", unsafe_allow_html=True)
            with col_b:
                st.markdown(f"**{comp_m_sel} {comp_a_sel}** (comparación)")
                def _diff(v_act, v_comp, suffix="", inverse=False):
                    diff = v_act - v_comp
                    color = "normal" if (diff >= 0) != inverse else "inverse"
                    return f"{diff:+.2f}{suffix}"
                st.metric("SLA",             f"{kpis_comp['global']['sla']:.3f}%",
                          delta=f"{kpis['global']['sla']-kpis_comp['global']['sla']:+.3f}%", delta_color="normal")
                st.metric("MTTR T1",         f"{kpis_comp['t1']['mttr']:.2f} h",
                          delta=f"{kpis['t1']['mttr']-kpis_comp['t1']['mttr']:+.2f}h", delta_color="inverse")
                st.metric("ACD",             f"{kpis_comp['t1']['acd']:.2f} h",
                          delta=f"{kpis['t1']['acd']-kpis_comp['t1']['acd']:+.2f}h", delta_color="inverse")
                st.metric("Eventos",         kpis_comp['global']['total_fallas'],
                          delta=f"{kpis['global']['total_fallas']-kpis_comp['global']['total_fallas']:+.0f}", delta_color="inverse")
                st.metric("Clientes Impact.", f"{kpis_comp['t1']['clientes']:,}",
                          delta=f"{kpis['t1']['clientes']-kpis_comp['t1']['clientes']:+.0f}", delta_color="inverse")
                st.metric("Impacto Acum.",   f"{kpis_comp['global']['db']/24:.2f} días",
                          delta=f"{(kpis['global']['db']-kpis_comp['global']['db'])/24:+.2f}d", delta_color="inverse")
                st.metric("MTBF",            f"{kpis_comp['global']['mtbf']:.1f} h",
                          delta=f"{kpis['global']['mtbf']-kpis_comp['global']['mtbf']:+.1f}h", delta_color="normal")
                st.metric("Fallas P1",       kpis_comp['global']['p1'],
                          delta=f"{kpis['global']['p1']-kpis_comp['global']['p1']:+.0f}", delta_color="inverse")
            st.divider()

        st.markdown("### 📍 SLA por Zona (Disponibilidad Geográfica)")
        st.caption("*Desglose de disponibilidad por cada nodo principal.*")
        zonas_activas = [z[0] for z in get_zonas()]
        sla_data      = [{"Zona": z, "SLA (%)": kpis['zonas_sla'].get(z, 100.0)} for z in zonas_activas]
        df_sla_geo    = pd.DataFrame(sla_data).sort_values("SLA (%)", ascending=True)
        sla_min       = df_sla_geo["SLA (%)"].min() if not df_sla_geo.empty else 98.0
        fig_sla = px.bar(df_sla_geo, x="SLA (%)", y="Zona", orientation='h',
                         text_auto='.3f', color="SLA (%)",
                         color_continuous_scale=["#ff2b2b","#ff9f43","#29b09d"],
                         title="Ranking de Estabilidad Geográfica")
        fig_sla.update_layout(xaxis=dict(range=[max(0, sla_min-2), 100.5]),
                               margin=dict(l=0,r=0,t=40,b=0), paper_bgcolor="rgba(0,0,0,0)", height=350)
        st.plotly_chart(fig_sla, use_container_width=True)

        st.divider()
        st.markdown("### 👥 Impacto Directo a Clientes (Datos Exactos)")
        st.info("💡 **Inteligencia de Datos:** Valores calculados con el histórico de eventos previos (CMDB Autónoma).")

        if kpis['t1']['fallas'] == 0:
            st.success("🟢 No hay incidentes masivos cerrados con afectación a clientes en este periodo.")
        else:
            t1a, t1b, t1c, t1d = st.columns(4)
            t1a.metric("Fallas Completas",       kpis['t1']['fallas'],
                       delta=_delta('t1','fallas',   fmt="{:+.0f}"), delta_color="inverse")
            t1b.metric("MTTR Real (Resolución)", f"{kpis['t1']['mttr']:.2f} horas",
                       delta=_delta('t1','mttr',     fmt="{:+.2f}", suffix="h"), delta_color="inverse")
            t1c.metric("ACD Real (Afectación)",  f"{kpis['t1']['acd']:.2f} horas",
                       delta=_delta('t1','acd',      fmt="{:+.2f}", suffix="h"), delta_color="inverse")
            t1d.metric("Clientes Afectados",     f"{kpis['t1']['clientes']:,}",
                       delta=_delta('t1','clientes', fmt="{:+.0f}"), delta_color="inverse")

        st.divider()
        dibujar_graficos(df_m)

t_idx += 1

# ─────────────────────────────────────────────
# TAB 1 — REGISTRAR EVENTO
# ─────────────────────────────────────────────
if role in ('admin', 'auditor'):
    with tabs[t_idx]:
        st.title("📝 Registrar Evento Operativo")
        fk = st.session_state.form_reset

        zonas_form   = [z[0] for z in get_zonas()]
        equipos_form = get_cat("cat_equipos")
        causas_form  = list(get_causas_con_flag().keys())
        servs_form   = get_cat("cat_servicios")

        cf, ccx = st.columns([2, 1], gap="large")
        with cf:
            with st.container(border=True):
                c_z1, c_z2 = st.columns([1,1])
                with c_z1:
                    z_f = st.selectbox("📍 Nodo Principal", zonas_form, key=f"z_{fk}")
                with c_z2:
                    st.markdown("<div style='margin-top:32px;'></div>", unsafe_allow_html=True)
                    ag = st.toggle("🚨 Falla General (Afecta todo el nodo)", value=True, key=f"ag_{fk}")

                sz    = st.text_input("📍 Sub-zona", value="General" if ag else "", key=f"sz_{fk}", disabled=ag)
                sz_db = "General" if ag else sz
                st.divider()

                c1f, c2f = st.columns(2)
                srv_f = c1f.selectbox("🌐 Servicio",  servs_form,              key=f"s_{fk}")
                cat_f = c2f.selectbox("🏢 Segmento",  list(DEFAULT_CATEGORIAS), key=f"c_{fk}")
                eq_f  = st.selectbox("🖥️ Equipo",     equipos_form,            key=f"e_{fk}")
                d_cl  = get_clientes_cmdb(z_f, eq_f)
                st.divider()

                if z_f == "San Salvador (Central)":
                    cl_f = 0; imp_pct = 0.0
                    st.number_input("👤 Clientes Afectados", value=0, disabled=True, key=f"cl_{fk}")
                    st.warning("⚠️ Centro de Datos: Equipos Core. No aplica conteo de ONTs.")
                elif cat_f == "Cliente Corporativo":
                    cl_f = 1; imp_pct = 0.0
                    st.number_input("👤 Clientes Afectados", value=1, disabled=True, key=f"cl_{fk}")
                    st.caption("🏢 Corporativo: fijado en 1 enlace.")
                elif cat_f == CAT_INTERNA:
                    cl_f = 0; imp_pct = 0.0
                    st.number_input("👤 Clientes Afectados", value=0, disabled=True, key=f"cl_{fk}")
                    st.caption("🔧 Interna: fijado en 0 clientes.")
                else:
                    cl_f    = st.number_input(f"👤 Clientes Afectados *(histórico CMDB: {d_cl})*", min_value=0, value=d_cl, step=1, key=f"cl_{fk}")
                    imp_pct = round((cl_f / d_cl) * 100, 2) if d_cl > 0 else 0.0

                st.divider()
                estado_ticket = st.radio("🚦 Estado del Evento", ["Cerrado (Falla Resuelta)", "Abierto (Falla en Curso)"], horizontal=True, key=f"estado_{fk}")
                es_abierto    = "Abierto" in estado_ticket

                ct1, ct2 = st.columns(2)
                with ct1:
                    fi    = st.date_input("📅 Fecha de Inicio", key=f"fi_{fk}")
                    hi_on = st.checkbox("🕒 Conozco la hora exacta de inicio", value=False, key=f"hi_on_{fk}")
                    if hi_on:
                        hi_val = st.time_input("Hora Exacta de Inicio", key=f"hi_val_{fk}")
                    else:
                        hi_val = None; st.info("ℹ️ Sin hora exacta → Evento Incompleto.")

                with ct2:
                    if es_abierto:
                        ff = fi; hf_val = None; hf_on = False
                        st.warning("🚨 El ticket se guardará como 'Falla en Curso'.")
                    else:
                        ff    = st.date_input("📅 Fecha de Cierre", key=f"ff_{fk}")
                        hf_on = st.checkbox("🕒 Conozco la hora exacta de cierre", value=False, key=f"hf_on_{fk}")
                        if hf_on:
                            hf_val = st.time_input("Hora Exacta de Cierre", key=f"hf_val_{fk}")
                        else:
                            hf_val = None; st.info("ℹ️ Sin hora exacta → Evento Incompleto.")

                dur = 0.0; conocimiento = "Parcial"
                if hi_on and hf_on and (not es_abierto) and hi_val and hf_val:
                    try:
                        dt_ini = datetime.combine(fi, hi_val)
                        dt_fin = datetime.combine(ff, hf_val)
                        if dt_fin > dt_ini:
                            dur = max(0.01, round((dt_fin - dt_ini).total_seconds() / 3600, 2))
                            conocimiento = "Total"
                    except Exception: pass

                st.write("")
                cr_f   = st.selectbox("🛠️ Causa Raíz", causas_form, key=f"cr_{fk}")
                desc_f = st.text_area("📝 Descripción del Evento", key=f"desc_{fk}")

                if st.button("💾 Guardar Registro", type="primary"):
                    err = False
                    if (not es_abierto) and hi_on and hf_on and hi_val and hf_val:
                        if fi > ff or (fi == ff and hi_val >= hf_val):
                            st.session_state.flash_msg  = "La fecha/hora de cierre no puede ser anterior al inicio."
                            st.session_state.flash_type = "error"; err = True

                    if not err:
                        with st.spinner("Guardando en Base de Datos…"):
                            try:
                                hi_db = hi_val if hi_on else datetime_time(0, 0)
                                idi   = SV_TZ.localize(datetime.combine(fi, hi_db))
                                idf   = None if (es_abierto or not hf_on or not hf_val) else SV_TZ.localize(datetime.combine(ff, hf_val))
                                estado_db = 'Abierto' if es_abierto else 'Cerrado'
                                with engine.begin() as conn:
                                    conn.execute(text("""
                                        INSERT INTO incidents (zona,subzona,afectacion_general,servicio,categoria,
                                          equipo_afectado,estado,inicio_incidente,fin_incidente,clientes_afectados,
                                          impacto_porcentaje,causa_raiz,descripcion,duracion_horas,conocimiento_tiempos)
                                        VALUES (:z,:sz,:ag,:s,:c,:e,:est,:idi,:idf,:cl,:imp,:cr,:d,:dur,:con)
                                    """), {"z":z_f,"sz":sz_db,"ag":ag,"s":srv_f,"c":cat_f,"e":eq_f,
                                           "est":estado_db,"idi":idi,"idf":idf,"cl":cl_f,"imp":imp_pct,
                                           "cr":cr_f,"d":desc_f,"dur":dur,"con":conocimiento})
                                    if z_f != "San Salvador (Central)":
                                        conn.execute(text("""
                                            INSERT INTO cmdb_nodos (zona,equipo,clientes,fuente) VALUES (:z,:e,:cl,'Manual')
                                            ON CONFLICT (zona,equipo) DO UPDATE SET clientes=EXCLUDED.clientes
                                            WHERE EXCLUDED.clientes > cmdb_nodos.clientes
                                        """), {"z":z_f,"e":eq_f,"cl":cl_f})
                                log_audit("INSERT", f"Falla ({estado_db}) en {z_f}")
                                _invalidar_cache()
                                st.session_state.form_reset += 1
                                st.session_state.flash_msg  = "✅ Registro guardado exitosamente."
                                st.session_state.flash_type = "success"; st.rerun()
                            except Exception as e:
                                st.session_state.flash_msg  = f"Error de BD: {e}"
                                st.session_state.flash_type = "error"; st.rerun()

        with ccx:
            st.markdown("#### 🕒 Registros Recientes")
            if not df_m.empty and 'id' in df_m.columns:
                for _, r in df_m.sort_values('id', ascending=False).head(6).iterrows():
                    with st.container(border=True):
                        ico = "🚨" if r.get('estado')=='Abierto' else ("🔧" if r.get('categoria')==CAT_INTERNA else "📡")
                        st.markdown(f"**{ico} {r.get('zona_completa', r.get('zona',''))}**")
                        st.caption(f"{str(r.get('causa_raiz',''))[:30]}… | 👥 {r.get('clientes_afectados',0)}")
            else:
                st.info("Sin registros en la tabla.")
    t_idx += 1

# ─────────────────────────────────────────────
# TAB 2 — HISTORIAL Y EDICIÓN
# ─────────────────────────────────────────────
if role in ('admin', 'auditor'):
    with tabs[t_idx]:
        st.markdown("### 🗂️ Historial, Auditoría y Edición")

        df_abiertos = df_m[df_m['estado'] == 'Abierto'] if ('estado' in df_m.columns and not df_m.empty) else pd.DataFrame()

        if not df_abiertos.empty:
            st.warning("🚨 Tienes fallas en curso. Puedes cerrarlas rápidamente aquí:")
            with st.expander("⚡ Cerrar Falla en Curso (Ticket Abierto)", expanded=True):
                with st.form("form_cerrar_ticket"):
                    def _fmt_inicio(ts):
                        try:
                            return ts.strftime('%d/%m/%Y %H:%M') if pd.notnull(ts) else 'Sin hora'
                        except Exception: return 'Sin hora'

                    t_opts = {
                        r['id']: f"ID: {r['id']} | Nodo: {r.get('zona_completa', r.get('zona',''))} | Inicio: {_fmt_inicio(r['inicio_incidente'])}"
                        for _, r in df_abiertos.iterrows()
                    }
                    sel_t  = st.selectbox("Selecciona la Falla a Cerrar", options=list(t_opts.keys()), format_func=lambda x: t_opts[x])
                    c_fc1, c_fc2 = st.columns(2)
                    f_fin  = c_fc1.date_input("📅 Fecha de Restablecimiento del Servicio")
                    h_fin  = c_fc2.time_input("🕒 Hora Exacta de Restablecimiento")

                    if st.form_submit_button("✅ Cerrar Ticket", type="primary"):
                        r_orig     = df_abiertos[df_abiertos['id'] == sel_t].iloc[0]
                        ini_dt     = r_orig['inicio_incidente']
                        fin_dt_val = SV_TZ.localize(datetime.combine(f_fin, h_fin))

                        if pd.isnull(ini_dt):
                            st.error("❌ Ticket sin hora de inicio. Edítalo manualmente en la tabla.")
                        elif fin_dt_val < ini_dt:
                            st.error("❌ La hora de cierre no puede ser anterior al inicio.")
                        else:
                            dur_h = max(0.01, round((fin_dt_val - ini_dt).total_seconds() / 3600, 2))
                            with engine.begin() as conn:
                                conn.execute(text("UPDATE incidents SET estado='Cerrado',fin_incidente=:f,duracion_horas=:d,conocimiento_tiempos='Total' WHERE id=:id"),
                                             {"f":fin_dt_val,"d":dur_h,"id":sel_t})
                            log_audit("CLOSE TICKET", f"Ticket ID {sel_t} cerrado.")
                            _invalidar_cache()
                            st.session_state.flash_msg  = "✅ Ticket cerrado correctamente."
                            st.session_state.flash_type = "success"; st.rerun()

        st.divider()
        papelera = st.toggle("🗑️ Explorar Papelera de Reciclaje")
        df_audit = load_and_enrich(fecha_ini, fecha_fin, papelera, "Todas", "Todos", "Todos", cache_version=_dv())

        if df_audit.empty:
            st.info("No hay datos en el servidor para el periodo.")
        else:
            c_s, c_pg = st.columns([4, 1])
            bq = c_s.text_input("🔎 Buscar:", placeholder="Causa, nodo, equipo…")
            df_d = (df_audit[df_audit.astype(str).apply(lambda x: x.str.contains(bq, case=False, na=False)).any(axis=1)].copy()
                    if bq else df_audit.copy())

            tot_p   = max(1, math.ceil(len(df_d) / 15))
            pg      = c_pg.number_input("Página", 1, tot_p, 1, key="p_bd")
            df_page = df_d.iloc[(pg-1)*15 : pg*15].copy()

            drop_cols = ['deleted_at','Severidad','zona_completa','es_externa','impacto_porcentaje']

            # ── Expander: ELIMINAR / RESTAURAR ──
            with st.expander(
                "🗑️ Mover Registros a la Papelera" if not papelera else "♻️ Restaurar o Eliminar Registros",
                expanded=False
            ):
                ids_disponibles = df_d['id'].tolist() if 'id' in df_d.columns else []

                def _label_id(x):
                    fila = df_d[df_d['id'] == x]
                    if fila.empty: return f"ID {x}"
                    zona_v  = fila['zona'].values[0]       if 'zona'       in fila.columns else ''
                    causa_v = fila['causa_raiz'].values[0] if 'causa_raiz' in fila.columns else ''
                    return f"ID {x} · {zona_v} · {str(causa_v)[:25]}"

                if papelera:
                    st.info("📌 Selecciona registros para **restaurarlos** o **eliminarlos permanentemente**.")
                    sel_ids = st.multiselect("Selecciona registros (por ID)", ids_disponibles, format_func=_label_id, key="sel_pap_ids")
                    if sel_ids:
                        col_rest, col_perm = st.columns(2)
                        if col_rest.button("♻️ Restaurar Seleccionados", type="primary", use_container_width=True):
                            with engine.begin() as conn:
                                for rid in sel_ids: conn.execute(text("UPDATE incidents SET deleted_at=NULL WHERE id=:id"), {"id":int(rid)})
                            log_audit("RESTORE", f"{len(sel_ids)} registro(s) restaurados.")
                            _invalidar_cache()
                            st.session_state.flash_msg = "♻️ Registros restaurados."; st.rerun()
                        if col_perm.button("🔥 Eliminar Permanentemente", use_container_width=True):
                            with engine.begin() as conn:
                                for rid in sel_ids: conn.execute(text("DELETE FROM incidents WHERE id=:id"), {"id":int(rid)})
                            log_audit("DELETE PERMANENT", f"{len(sel_ids)} eliminados.")
                            _invalidar_cache()
                            st.session_state.flash_msg = "🔥 Eliminados permanentemente."; st.rerun()
                    else:
                        st.caption("Selecciona al menos un registro.")
                else:
                    st.info("📌 Selecciona registros para moverlos a la papelera (reversible).")
                    sel_ids = st.multiselect("Selecciona registros (por ID)", ids_disponibles, format_func=_label_id, key="sel_del_ids")
                    if sel_ids:
                        if st.button("🗑️ Mover a Papelera", type="primary", use_container_width=True):
                            with engine.begin() as conn:
                                for rid in sel_ids: conn.execute(text("UPDATE incidents SET deleted_at=CURRENT_TIMESTAMP WHERE id=:id"), {"id":int(rid)})
                            log_audit("DELETE (SOFT)", f"{len(sel_ids)} movidos a papelera.")
                            _invalidar_cache()
                            st.session_state.flash_msg = "🗑️ Registros movidos a la papelera."; st.rerun()
                    else:
                        st.caption("Selecciona al menos un registro.")

            st.divider()

            # ── EDICIÓN DIRECTA DESDE LA TABLA ──
            st.markdown("#### ✏️ Edición Directa de Registros")
            st.info("⚠️ Edita las celdas directamente y presiona **💾 Guardar Ediciones Manuales** para confirmar.")

            ref_df = df_page.drop(columns=drop_cols, errors='ignore').reset_index(drop=True)
            ed_df  = st.data_editor(
                ref_df,
                column_config={
                    "id":               st.column_config.NumberColumn("🆔 ID", disabled=True, width="small"),
                    "zona":             st.column_config.TextColumn("📍 Zona"),
                    "estado":           st.column_config.SelectboxColumn("🚦 Estado", options=["Cerrado","Abierto"]),
                    "inicio_incidente": st.column_config.DatetimeColumn("🕐 Inicio", format="YYYY-MM-DD HH:mm"),
                    "fin_incidente":    st.column_config.DatetimeColumn("🕑 Fin",    format="YYYY-MM-DD HH:mm"),
                    "clientes_afectados": st.column_config.NumberColumn("👥 Clientes"),
                    "duracion_horas":   st.column_config.NumberColumn("⏱️ Horas",  format="%.2f"),
                    "descripcion":      st.column_config.TextColumn("📝 Descripción", width="large"),
                },
                use_container_width=True, hide_index=True,
                key="editor_incidentes_v12"
            )

            ref_stripped = _strip_tz_df(ref_df.reset_index(drop=True))
            ed_stripped  = _strip_tz_df(ed_df.reset_index(drop=True))
            h_cam        = not ref_stripped.equals(ed_stripped)

            if h_cam:
                if st.button("💾 Guardar Ediciones Manuales", type="primary", use_container_width=True):
                    filas_invalidas = [i for i in range(len(ed_df))
                                       if not ref_stripped.iloc[i].equals(ed_stripped.iloc[i])
                                       and pd.isnull(ed_df.iloc[i].get('inicio_incidente'))]
                    if filas_invalidas:
                        st.session_state.flash_msg  = "❌ 'Fecha de Inicio' es obligatoria en filas editadas."
                        st.session_state.flash_type = "error"; st.rerun()
                    else:
                        with engine.begin() as conn:
                            for i in range(len(ed_df)):
                                if ref_stripped.iloc[i].equals(ed_stripped.iloc[i]): continue
                                r = ed_df.iloc[i]
                                try:
                                    ini_dt = pd.to_datetime(r.get('inicio_incidente'))
                                    if pd.isnull(r.get('fin_incidente')):
                                        fin_dt_sql=None; dur_u=0.0; con_u="Parcial"; est_u="Abierto"
                                    else:
                                        fin_dt=pd.to_datetime(r.get('fin_incidente')); fin_dt_sql=fin_dt
                                        dur_u=max(0.01, round((fin_dt-ini_dt).total_seconds()/3600,2))
                                        con_u="Total"; est_u="Cerrado"
                                except Exception:
                                    fin_dt_sql=None; dur_u=0.0; con_u="Parcial"; est_u="Abierto"
                                try:
                                    cl_val = r.get('clientes_afectados',0)
                                    cl_val = int(float(cl_val)) if pd.notnull(cl_val) and str(cl_val).strip()!='' else 0
                                except Exception: cl_val = 0
                                row_id = r.get('id')
                                if pd.isnull(row_id): continue
                                conn.execute(text("""
                                    UPDATE incidents SET zona=:z,subzona=:sz,afectacion_general=:ag,
                                        servicio=:s,categoria=:c,equipo_afectado=:e,estado=:est,
                                        inicio_incidente=:idi,fin_incidente=:idf,clientes_afectados=:cl,
                                        causa_raiz=:cr,descripcion=:d,duracion_horas=:dur,
                                        conocimiento_tiempos=:con WHERE id=:id
                                """), {"z":str(r.get('zona','')),"sz":str(r.get('subzona','')),"ag":bool(r.get('afectacion_general',True)),
                                       "s":str(r.get('servicio','')),"c":str(r.get('categoria','')),"e":str(r.get('equipo_afectado','')),
                                       "est":est_u,"idi":ini_dt,"idf":fin_dt_sql,"cl":cl_val,
                                       "cr":str(r.get('causa_raiz','')),"d":str(r.get('descripcion','') or ''),
                                       "dur":dur_u,"con":con_u,"id":int(row_id)})
                        log_audit("UPDATE", "Edición masiva en tabla.")
                        _invalidar_cache()
                        st.session_state.flash_msg  = "💾 Cambios guardados correctamente."
                        st.session_state.flash_type = "success"; st.rerun()

            st.divider()
            # Descargas: CSV y Excel
            dl_csv, dl_excel = st.columns(2)
            with dl_csv:
                st.download_button(
                    "📥 Descargar CSV",
                    df_d.drop(columns=drop_cols, errors='ignore').to_csv(index=False).encode(),
                    f"NOC_Export_{fecha_ini}_{fecha_fin}.csv", "text/csv",
                    use_container_width=True
                )
            with dl_excel:
                kpis_excel = calc_kpis(df_m, fecha_ini, fecha_fin)
                st.download_button(
                    "📊 Descargar Excel (con formato)",
                    generar_excel(df_audit, kpis_excel, label_periodo),
                    f"NOC_{m_sel}_{a_sel}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    t_idx += 1

# ─────────────────────────────────────────────
# TAB 3 — CONFIGURACIÓN
# ─────────────────────────────────────────────
if role == 'admin' and len(tabs) > t_idx:
    with tabs[t_idx]:
        st.markdown("### ⚙️ Configuración del Sistema")
        st.caption("Administra catálogos y usuarios. Los cambios se reflejan inmediatamente en la UI.")

        t_zonas, t_equipos, t_causas, t_usuarios = st.tabs(["🗺️ Zonas", "🖥️ Equipos", "🛠️ Causas", "👤 Usuarios y Accesos"])

        with t_zonas:
            st.markdown("#### Gestión de Zonas / Nodos")
            for z_nombre, z_lat, z_lon in get_zonas():
                c_zn, c_zdel = st.columns([5, 1])
                c_zn.text(f"📍 {z_nombre}  (Lat: {z_lat:.4f}, Lon: {z_lon:.4f})")
                if c_zdel.button("🗑️ Eliminar", key=f"del_z_{z_nombre}"):
                    try:
                        with engine.begin() as conn: conn.execute(text("DELETE FROM cat_zonas WHERE nombre=:n"), {"n": z_nombre})
                        clear_catalog_cache(); st.session_state.flash_msg = "🗑️ Zona eliminada."; st.rerun()
                    except Exception:
                        st.session_state.flash_msg = "❌ La zona no puede eliminarse."; st.session_state.flash_type = "error"; st.rerun()
            st.divider()
            with st.form("form_add_zona", clear_on_submit=True):
                st.markdown("**➕ Agregar Nueva Zona**")
                nz = st.text_input("Nombre del Nodo Principal")
                cla, clo = st.columns(2)
                nlat = cla.number_input("Latitud",  value=13.6929, format="%.4f")
                nlon = clo.number_input("Longitud", value=-89.2182, format="%.4f")
                if st.form_submit_button("Agregar Zona") and nz:
                    try:
                        with engine.begin() as conn: conn.execute(text("INSERT INTO cat_zonas (nombre,lat,lon) VALUES (:n,:la,:lo)"), {"n":nz,"la":nlat,"lo":nlon})
                        clear_catalog_cache(); st.session_state.flash_msg = "✅ Zona agregada."; st.rerun()
                    except Exception: st.toast("Error: Zona duplicada.", icon="❌")

        with t_equipos:
            st.markdown("#### Gestión de Equipos de Red")
            for eq in get_cat("cat_equipos"):
                c_en, c_edel = st.columns([5,1])
                c_en.text(f"🖥️ {eq}")
                if c_edel.button("🗑️ Eliminar", key=f"del_eq_{eq}"):
                    try:
                        with engine.begin() as conn: conn.execute(text("DELETE FROM cat_equipos WHERE nombre=:n"), {"n": eq})
                        clear_catalog_cache(); st.session_state.flash_msg = "🗑️ Equipo eliminado."; st.rerun()
                    except Exception:
                        st.session_state.flash_msg = "❌ El equipo está en uso."; st.session_state.flash_type = "error"; st.rerun()
            st.divider()
            with st.form("form_add_eq", clear_on_submit=True):
                st.markdown("**➕ Agregar Nuevo Equipo**")
                ne = st.text_input("Nombre del dispositivo/equipo")
                if st.form_submit_button("Agregar Equipo") and ne:
                    try:
                        with engine.begin() as conn: conn.execute(text("INSERT INTO cat_equipos (nombre) VALUES (:n)"), {"n": ne})
                        clear_catalog_cache(); st.session_state.flash_msg = "✅ Equipo agregado."; st.rerun()
                    except Exception: st.toast("Error: Equipo duplicado.", icon="❌")

        with t_causas:
            st.markdown("#### Gestión de Causas Raíz")
            for causa, ext in get_causas_con_flag().items():
                c_cn, c_ct, c_cdel = st.columns([4,2,1])
                c_cn.text(f"🛠️ {causa}")
                c_ct.caption("Externa (Fuerza Mayor)" if ext else "Interna (NOC)")
                if c_cdel.button("🗑️ Eliminar", key=f"del_ca_{causa}"):
                    try:
                        with engine.begin() as conn: conn.execute(text("DELETE FROM cat_causas WHERE nombre=:n"), {"n": causa})
                        clear_catalog_cache(); st.session_state.flash_msg = "🗑️ Causa eliminada."; st.rerun()
                    except Exception:
                        st.session_state.flash_msg = "❌ La causa está en uso."; st.session_state.flash_type = "error"; st.rerun()
            st.divider()
            with st.form("form_add_causa", clear_on_submit=True):
                st.markdown("**➕ Agregar Nueva Causa**")
                nc = st.text_input("Descripción de la causa")
                nce = st.checkbox("¿Es un factor Externo (Terceros, Clima, etc)?")
                if st.form_submit_button("Agregar Causa") and nc:
                    try:
                        with engine.begin() as conn: conn.execute(text("INSERT INTO cat_causas (nombre,es_externa) VALUES (:n,:e)"), {"n":nc,"e":nce})
                        clear_catalog_cache(); st.session_state.flash_msg = "✅ Causa agregada."; st.rerun()
                    except Exception: st.toast("Error: Causa duplicada.", icon="❌")

        with t_usuarios:
            st.markdown("#### Control de Accesos y Contraseñas")
            st.info("🔒 **Política:** Mínimo 8 caracteres · Mayúscula · Minúscula · Número · Carácter especial (!@#$%...)")
            cu, clg = st.columns([1, 2], gap="large")
            with cu:
                with st.form("form_u", clear_on_submit=True):
                    st.markdown("**Crear Usuario**")
                    nu   = st.text_input("Usuario")
                    np_u = st.text_input("Contraseña", type="password")
                    nrl  = st.selectbox("Rol Asignado", ["viewer", "auditor", "admin"])
                    if st.form_submit_button("Crear Cuenta") and nu and np_u:
                        pw_error = validar_password(np_u)
                        if pw_error:
                            st.error(f"❌ {pw_error}")
                        else:
                            try:
                                with engine.begin() as conn:
                                    conn.execute(text("INSERT INTO users (username,password_hash,role) VALUES (:u,:h,:r)"), {"u":nu,"h":hash_pw(np_u),"r":nrl})
                                st.session_state.flash_msg = "✅ Usuario creado."; st.rerun()
                            except Exception as e:
                                if "unique" in str(e).lower() or "duplicate" in str(e).lower():
                                    st.toast("❌ Ese usuario ya existe.", icon="❌")
                                else:
                                    st.toast(f"❌ Error: {str(e)}", icon="❌")

                with st.form("form_reset_pw", clear_on_submit=True):
                    st.markdown("**Restablecer Contraseña**")
                    u_reset = st.text_input("Usuario exacto")
                    p_reset = st.text_input("Nueva Contraseña", type="password")
                    if st.form_submit_button("Restablecer") and u_reset and p_reset:
                        pw_error = validar_password(p_reset)
                        if pw_error:
                            st.error(f"❌ {pw_error}")
                        else:
                            try:
                                with engine.begin() as conn:
                                    res = conn.execute(text("UPDATE users SET password_hash=:h,failed_attempts=0,locked_until=NULL WHERE username=:u"), {"h":hash_pw(p_reset),"u":u_reset})
                                    if res.rowcount > 0:
                                        st.session_state.flash_msg = "✅ Contraseña actualizada."; st.rerun()
                                    else:
                                        st.toast("❌ Usuario no encontrado.", icon="❌")
                            except Exception as e: st.toast(f"Error: {e}")

            with clg:
                try:
                    with engine.connect() as conn:
                        df_usrs = pd.read_sql(text("SELECT id,username,role,is_banned,failed_attempts FROM users"), conn)
                    # FIX: fillna en columnas numéricas antes de cualquier operación
                    df_usrs['failed_attempts'] = df_usrs['failed_attempts'].fillna(0).astype(int)
                    df_usrs['is_banned']        = df_usrs['is_banned'].fillna(False)

                    df_admin_row = df_usrs[df_usrs['username'] == SUPERADMIN_USERNAME]
                    df_otros     = df_usrs[df_usrs['username'] != SUPERADMIN_USERNAME].reset_index(drop=True)

                    if not df_admin_row.empty:
                        st.markdown(f"🛡️ **Super Admin protegido** (no editable ni eliminable)")
                        st.dataframe(df_admin_row[['username','role','is_banned','failed_attempts']].rename(columns={
                            'username':'Usuario','role':'Rol','is_banned':'Baneado','failed_attempts':'Intentos'}),
                            hide_index=True, use_container_width=True)

                    if not df_otros.empty:
                        df_otros_edit = df_otros.copy()
                        df_otros_edit.insert(0, "Sel", False)

                        ed_usrs = st.data_editor(
                            df_otros_edit,
                            column_config={
                                "Sel":   st.column_config.CheckboxColumn("✔", default=False),
                                "id":    None,
                                "username": "Usuario",
                                # FIX #9: SelectboxColumn evita roles inválidos y falsos positivos en comparación
                                "role":  st.column_config.SelectboxColumn("Rol", options=["viewer","auditor","admin"]),
                                "is_banned":       "Baneado",
                                "failed_attempts": "Intentos Fallidos",
                            },
                            use_container_width=True, hide_index=True,
                            key="editor_usuarios_v12"
                        )
                        filas_del = ed_usrs[ed_usrs["Sel"] == True]

                        # FIX #9: comparar con .equals() — no fillna('') que convierte bool a str
                        orig_cmp = df_otros_edit.drop(columns=['Sel'], errors='ignore').reset_index(drop=True)
                        edit_cmp = ed_usrs.drop(columns=['Sel'], errors='ignore').reset_index(drop=True)
                        hay_cambios = not orig_cmp.equals(edit_cmp)

                        u1c, u2c = st.columns(2)
                        if not filas_del.empty and u1c.button("🗑️ Eliminar Usuario", use_container_width=True):
                            if SUPERADMIN_USERNAME in filas_del['username'].values:
                                st.error(f"❌ No se puede eliminar la cuenta {SUPERADMIN_USERNAME} raíz.")
                            elif st.session_state.username in filas_del['username'].values:
                                st.session_state.flash_msg = "❌ No puedes eliminar tu propia cuenta."
                                st.session_state.flash_type = "error"; st.rerun()
                            else:
                                with engine.begin() as conn:
                                    for rid in filas_del['id']:
                                        if pd.notnull(rid):
                                            conn.execute(text("DELETE FROM users WHERE id=:id"), {"id": int(rid)})
                                st.session_state.flash_msg = "🗑️ Usuario eliminado."; st.rerun()

                        if hay_cambios and u2c.button("💾 Guardar Permisos", type="primary", use_container_width=True):
                            with engine.begin() as conn:
                                for i, er in ed_usrs.iterrows():
                                    orig = df_otros_edit.drop(columns=['Sel'], errors='ignore').iloc[i]
                                    if str(er.get('username','')) == SUPERADMIN_USERNAME: continue
                                    if not orig.equals(er.drop('Sel', errors='ignore')):
                                        er_id = er.get('id')
                                        if pd.isnull(er_id): continue
                                        # FIX: failed_attempts nunca None
                                        fa_val = int(er.get('failed_attempts', 0) or 0)
                                        conn.execute(text("UPDATE users SET role=:r,is_banned=:b,failed_attempts=:f WHERE id=:id"),
                                                     {"r":str(er.get('role','viewer')),"b":bool(er.get('is_banned',False)),
                                                      "f":fa_val,"id":int(er_id)})
                            st.session_state.flash_msg = "💾 Permisos guardados."; st.rerun()
                    else:
                        st.info("No hay otros usuarios registrados.")

                except Exception as e:
                    st.error(str(e))

            st.divider()
            st.markdown("##### 📜 Registro del Sistema (Audit Log)")
            try:
                with engine.connect() as conn:
                    logs = pd.read_sql(text("SELECT timestamp AS Fecha,username AS Usuario,action AS Accion,details AS Detalles FROM audit_logs ORDER BY id DESC"), conn)
                t_lp  = max(1, math.ceil(len(logs)/10))
                p_log = st.number_input("Página de log", 1, t_lp, 1)
                st.dataframe(logs.iloc[(p_log-1)*10 : p_log*10], use_container_width=True, hide_index=True)
            except Exception as e:
                st.warning(str(e))
